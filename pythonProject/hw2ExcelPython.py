import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle
from glob import glob
from pathlib import Path

# Analysis 2 – Wrangling csv files
# Get current working directory
cwd = os.getcwd()
print(cwd)
print(type(cwd))

# List all files and directories in data directory
print(os.listdir("./logs"))

# Create a Path object corresponding to the new directory name
logs_dir = Path.cwd() / 'logs'
print(logs_dir)

# Make the new directory
logs_dir.mkdir(exist_ok=True)
print(logs_dir.is_dir())

# Find all csv files in logs directory
csv_files = list(logs_dir.glob('*.csv'))

# Create a Pandas Excel writer using openpyxl as the engine
BCM_file = logs_dir / 'BCM.xlsx'
with pd.ExcelWriter(BCM_file, engine='openpyxl') as writer:
    for csv_file in csv_files:
        # Read each CSV file into a DataFrame
        df = pd.read_csv(csv_file)

        # Add headers
        headers = ['Datetime', 'Scale', 'Temperature']
        df.columns = headers

        # Convert the 'Datetime' column to datetime objects
        df['Datetime'] = pd.to_datetime(df['Datetime'])

        # Use the name of the csv file as the sheet name
        sheet_name = csv_file.stem

        # Write the DataFrame to an Excel sheet
        df.to_excel(writer, sheet_name=sheet_name, index=False)

print(f"All CSV files have been copied to {BCM_file}")

# Reopen excel file
wb = load_workbook(BCM_file)

# Define style for formatting cells
highlight = NamedStyle(name="highlight")
highlight.number_format = 'yyyy-mm-dd hh:mm:ss'

temperature_style = NamedStyle(name="temperature")
temperature_style.number_format = '0.00'

for sheet in wb.worksheets:
    if sheet.title != 'Sheet':
        # Get the max row number
        max_row = sheet.max_row

        # Add labels for min, max, average in G2:G4
        sheet['G2'] = 'min_temp'
        sheet['G3'] = 'max_temp'
        sheet['G4'] = 'mean_temp'

        # Add formulas for min, max, average in H2:H4
        sheet['H2'] = f"=MIN(C2:C{max_row})"
        sheet['H3'] = f"=MAX(C2:C{max_row})"
        sheet['H4'] = f"=AVERAGE(C2:C{max_row})"

        # Format cells in column H
        sheet['H2'].style = highlight
        sheet['H3'].style = highlight
        sheet['H4'].style = highlight

        # Add labels for min, max datetime in G6:G7
        sheet['G6'] = 'min_date'
        sheet['G7'] = 'max_date'

        # Add formulas for min, max datetime in H6:H7
        sheet['H6'] = f"=MIN(A2:A{max_row})"
        sheet['H7'] = f"=MAX(A2:A{max_row})"

        # Format cells in column H
        sheet['H6'].style = highlight
        sheet['H7'].style = highlight

# Adjust column widths
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter  # Get the column name
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

# Save the modified Excel file
wb.save(BCM_file)

print(f"Formulas and column adjustments have been added to {BCM_file}")

# Apply number format directly to cells containing temperature calculations
for sheet in wb.worksheets:
    if sheet.title != 'Sheet':
        sheet['H2'].number_format = '0.00'
        sheet['H3'].number_format = '0.00'
        sheet['H4'].number_format = '0.00'

# Save the workbook to finalize number format changes
wb.save(BCM_file)

print(f"Number format applied to temperature calculations in {BCM_file}")














